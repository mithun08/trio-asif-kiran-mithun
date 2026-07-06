from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from matcher.config import OCRConfig
from matcher.models.consultant import Consultant, Skill
from matcher.models.errors import IngestionError
from matcher.models.role import RequiredSkill, Role
from matcher.pipeline.normalise import normalise_location
from matcher.pipeline.store import hash_file

_PROFICIENCY_MAP: dict[str, int] = {
    "expert": 5,
    "working": 3,
    "experienced": 3,
    "proficient": 3,
    "beginner": 1,
    "learning": 1,
}

_SUPPLY_STATE_MAP: dict[str, str] = {
    "Beach": "beach",
    "Rolling Off": "rolling_off",
    "New Joiners": "new_joiner",
}

_SECTION_PATTERN = re.compile(
    r"^## (Project|Client|Beach) feedback",
    re.MULTILINE | re.IGNORECASE,
)
_EMAIL_KEY_PATTERN = re.compile(r"\*\*Email \(key\):\*\*\s*(\S+)", re.IGNORECASE)
_SECTION_SOURCE_MAP = {
    "project": "project",
    "client": "client",
    "beach": "beach",
}


def _parse_date(val: object) -> date | None:
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    try:
        return date.fromisoformat(str(val).strip())
    except (ValueError, TypeError):
        return None


def _parse_required_skill(token: str) -> RequiredSkill:
    token = token.strip()
    match = re.match(r"^(.+?)\s*\((\w+)\)$", token)
    if match:
        name, prof_text = match.group(1).strip(), match.group(2).casefold()
        proficiency = _PROFICIENCY_MAP.get(prof_text)
        return RequiredSkill(name=name, required_proficiency=proficiency)
    return RequiredSkill(name=token)


def _build_header_map(ws: Any, header_row: int = 2) -> dict[str, int]:
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    return {str(cell): idx for idx, cell in enumerate(row) if cell is not None}


def ingest_roles(xlsx_path: Path) -> list[Role]:
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        raise IngestionError(xlsx_path, "workbook unreadable") from exc
    try:
        ws = wb["Open Roles"]
    except KeyError:
        raise IngestionError(xlsx_path, "missing sheet 'Open Roles'")
    headers = _build_header_map(ws, header_row=2)

    required_columns = {"Role ID", "Title", "Required Skills", "Location"}
    missing = required_columns - headers.keys()
    if missing:
        raise IngestionError(xlsx_path, f"Open Roles sheet missing columns: {missing}")

    roles: list[Role] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        role_id = row[headers["Role ID"]]
        if not role_id:
            continue

        skills_raw = row[headers.get("Required Skills", -1)] or ""
        required_skills = [
            _parse_required_skill(t) for t in str(skills_raw).split(";") if t.strip()
        ]

        loc_raw = str(row[headers.get("Location", -1)] or "")
        locations = [normalise_location(p.strip()) for p in loc_raw.split("/") if p.strip()]

        co_located = str(row[headers.get("Co-location", -1)] or "").strip().casefold() == "yes"
        start_date = _parse_date(row[headers.get("Start", -1)])
        description = str(row[headers.get("Notes / Constraints", -1)] or "")
        sector_idx = headers.get("Sector")
        sector = str(row[sector_idx] or "") if sector_idx is not None else ""

        roles.append(
            Role(
                id=str(role_id),
                title=str(row[headers["Title"]] or ""),
                description=description,
                required_skills=required_skills,
                locations=locations,
                co_located=co_located,
                start_date=start_date,
                sector=sector,
            )
        )
    return roles


def _parse_skills(raw: object) -> list[Skill]:
    if not raw:
        return []
    return [
        Skill(name=s.strip(), proficiency=3, years_experience=0.0)
        for s in str(raw).split(",")
        if s.strip()
    ]


def ingest_consultants_from_workbook(xlsx_path: Path) -> list[Consultant]:
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        raise IngestionError(xlsx_path, "workbook unreadable") from exc
    result: list[Consultant] = []

    for sheet_name, supply_state in _SUPPLY_STATE_MAP.items():
        try:
            ws = wb[sheet_name]
        except KeyError:
            raise IngestionError(xlsx_path, f"missing sheet '{sheet_name}'")
        headers = _build_header_map(ws, header_row=2)

        required_columns = {"Name", "Email", "Grade", "Location"}
        missing = required_columns - headers.keys()
        if missing:
            raise IngestionError(xlsx_path, f"{sheet_name} sheet missing columns: {missing}")

        skills_col = next(
            (k for k in headers if k.startswith("Key Skills")),
            None,
        )

        for row in ws.iter_rows(min_row=3, values_only=True):
            name = row[headers["Name"]]
            if not name:
                continue

            skills_raw = row[headers[skills_col]] if skills_col else None

            available_from: date | None = None
            rolloff_confidence = "high"
            days_on_beach = 0

            if supply_state == "rolling_off":
                available_from = _parse_date(row[headers.get("Roll-off Date", -1)])
                rolloff_confidence = str(row[headers.get("Confidence", -1)] or "high").casefold()
            elif supply_state == "new_joiner":
                available_from = _parse_date(row[headers.get("Join Date", -1)])
            elif supply_state == "beach":
                dob_idx = headers.get("Days on Beach")
                dob_raw = row[dob_idx] if dob_idx is not None else None
                days_on_beach = (
                    int(dob_raw) if isinstance(dob_raw, (int, float)) and dob_raw >= 0 else 0
                )

            result.append(
                Consultant(
                    name=str(name),
                    email=str(row[headers["Email"]] or ""),
                    grade=str(row[headers["Grade"]] or ""),
                    location=str(row[headers["Location"]] or ""),
                    skills=_parse_skills(skills_raw),
                    raw_profile_text=str(row[headers.get("Notes", -1)] or ""),
                    available_from=available_from,
                    supply_state=supply_state,  # type: ignore[arg-type]
                    rolloff_confidence=rolloff_confidence,  # type: ignore[arg-type]
                    days_on_beach=days_on_beach,
                )
            )

    return result


def _derive_name_from_pdf_stem(stem: str) -> str:
    stem = stem.removesuffix("_pp")
    parts = stem.split("_")
    return " ".join(part.capitalize() for part in parts if part)


def _extract_pdf_text(
    pdf_path: Path, ocr_config: OCRConfig | None = None
) -> tuple[str, list[str], float]:
    from docling.document_converter import DocumentConverter

    _ocr = ocr_config if ocr_config is not None else OCRConfig()

    try:
        converter = DocumentConverter()
        result = converter.convert(str(pdf_path))
        raw_text = result.document.export_to_text()
    except Exception:
        return "", ["profile_pdf_unreadable"], 0.7

    if len(raw_text) < _ocr.text_floor_chars:
        if _ocr.enabled:
            try:
                from docling.document_converter import DocumentConverter as _DC

                ocr_converter = _DC()
                ocr_result = ocr_converter.convert(str(pdf_path))
                ocr_text = ocr_result.document.export_to_text()
                if ocr_text.strip():
                    return ocr_text, ["profile_pdf_ocr_used"], _ocr.confidence_floor
            except Exception:
                pass
        return "", ["profile_pdf_low_confidence"], 0.4

    return raw_text, [], 1.0


def _ocr_config_fingerprint(ocr_config: OCRConfig | None) -> str:
    cfg = ocr_config if ocr_config is not None else OCRConfig()
    return f"{cfg.enabled}:{cfg.text_floor_chars}:{cfg.confidence_floor}"


def _extract_pdf_text_cached(
    pdf_path: Path,
    ocr_config: OCRConfig | None,
    cache: dict[str, dict[str, Any]] | None,
) -> tuple[str, list[str], float]:
    if cache is None:
        return _extract_pdf_text(pdf_path, ocr_config)

    validity_key = f"{hash_file(pdf_path)}|{_ocr_config_fingerprint(ocr_config)}"
    cached = cache.get(pdf_path.name)
    if cached is not None and cached.get("validity_key") == validity_key:
        return cached["raw_text"], list(cached["data_gaps"]), cached["confidence_factor"]

    raw_text, data_gaps, confidence_factor = _extract_pdf_text(pdf_path, ocr_config)
    if raw_text:
        cache[pdf_path.name] = {
            "validity_key": validity_key,
            "raw_text": raw_text,
            "data_gaps": data_gaps,
            "confidence_factor": confidence_factor,
        }
    return raw_text, data_gaps, confidence_factor


def ingest_consultants(
    profiles_dir: Path,
    workbook_consultants: list[Consultant],
    ocr_config: OCRConfig | None = None,
    cache: dict[str, dict[str, Any]] | None = None,
) -> list[Consultant]:
    import logging

    if not profiles_dir.exists():
        logging.getLogger(__name__).warning("profiles_dir not found: %s", profiles_dir)
        return workbook_consultants

    consultant_by_name: dict[str, Consultant] = {c.name.casefold(): c for c in workbook_consultants}
    matched_emails: set[str] = set()

    updated: dict[str, Consultant] = {c.email.casefold(): c for c in workbook_consultants}

    for pdf_path in sorted(profiles_dir.glob("*.pdf")):
        derived_name = _derive_name_from_pdf_stem(pdf_path.stem)
        matched = consultant_by_name.get(derived_name.casefold())

        if matched is None:
            logging.getLogger(__name__).warning("No workbook match for PDF: %s", pdf_path.name)
            continue

        raw_text, data_gaps_update, confidence_factor = _extract_pdf_text_cached(
            pdf_path, ocr_config, cache
        )

        current = updated[matched.email.casefold()]
        new_gaps = [*current.data_gaps, *data_gaps_update]
        new_confidence = current.data_confidence * confidence_factor

        update: dict[str, Any] = {"data_gaps": new_gaps, "data_confidence": new_confidence}
        if raw_text:
            update["raw_profile_text"] = raw_text
            matched_emails.add(matched.email.casefold())

        updated[matched.email.casefold()] = current.model_copy(update=update)

    result = []
    for consultant in workbook_consultants:
        final = updated[consultant.email.casefold()]
        if consultant.email.casefold() not in matched_emails and not final.raw_profile_text:
            existing_gaps = [g for g in final.data_gaps if g != "profile_pdf_unmatched"]
            final = final.model_copy(
                update={"data_gaps": [*existing_gaps, "profile_pdf_unmatched"]}
            )
        result.append(final)

    return result


def _parse_feedback_sections(content: str) -> dict[str, str]:
    sections: dict[str, str] = {}
    parts = _SECTION_PATTERN.split(content)

    i = 1
    while i + 1 < len(parts):
        section_name = parts[i].strip().casefold()
        section_text = parts[i + 1].strip()
        source_key = _SECTION_SOURCE_MAP.get(section_name)
        if source_key and section_text:
            sections[source_key] = section_text
        i += 2

    return sections


def ingest_feedback(feedback_dir: Path, consultants: list[Consultant]) -> list[Consultant]:
    import logging

    if not feedback_dir.exists():
        logging.getLogger(__name__).warning("feedback_dir not found: %s", feedback_dir)
        return consultants

    consultant_by_email: dict[str, Consultant] = {c.email.casefold(): c for c in consultants}
    updated: dict[str, Consultant] = dict(consultant_by_email)

    for md_path in sorted(feedback_dir.glob("*.md")):
        content = md_path.read_text(encoding="utf-8")
        email_match = _EMAIL_KEY_PATTERN.search(content)

        if email_match is None:
            logging.getLogger(__name__).warning("No email key in %s", md_path.name)
            continue

        email_key = email_match.group(1).casefold()
        if email_key not in consultant_by_email:
            logging.getLogger(__name__).warning("Orphan feedback file: %s", md_path.name)
            continue

        feedback_text = _parse_feedback_sections(content)
        if not feedback_text:
            continue

        current = updated[email_key]
        merged_feedback = {**current.feedback_text, **feedback_text}
        update_fb: dict[str, Any] = {"feedback_text": merged_feedback}
        updated[email_key] = current.model_copy(update=update_fb)

    return [updated.get(c.email.casefold(), c) for c in consultants]
