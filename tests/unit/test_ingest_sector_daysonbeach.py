from __future__ import annotations

from pathlib import Path

import openpyxl

from matcher.pipeline.ingest import ingest_consultants_from_workbook, ingest_roles


def _make_roles_workbook(tmp_path: Path, with_sector: bool = True) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Open Roles"
    headers = ["Role ID", "Title", "Required Skills", "Location"]
    if with_sector:
        headers.append("Sector")
    ws.append([""] * len(headers))
    ws.append(headers)
    row = ["R1", "Dev", "Python", "London"]
    if with_sector:
        row.append("FinTech")
    ws.append(row)
    path = tmp_path / "roles.xlsx"
    wb.save(path)
    return path


def _make_consultants_workbook(
    tmp_path: Path, with_days_on_beach: bool = True, include_rolling_off: bool = False
) -> Path:
    wb = openpyxl.Workbook()

    beach_ws = wb.active
    beach_ws.title = "Beach"
    beach_headers = ["Name", "Email", "Grade", "Location", "Key Skills"]
    if with_days_on_beach:
        beach_headers.append("Days on Beach")
    beach_ws.append([""] * len(beach_headers))
    beach_ws.append(beach_headers)
    beach_row = ["Alice", "alice@x.com", "Senior", "London", "Python"]
    if with_days_on_beach:
        beach_row.append(45)
    beach_ws.append(beach_row)

    ro_ws = wb.create_sheet("Rolling Off")
    ro_headers = ["Name", "Email", "Grade", "Location", "Key Skills", "Roll-off Date", "Confidence"]
    ro_ws.append([""] * len(ro_headers))
    ro_ws.append(ro_headers)
    if include_rolling_off:
        ro_ws.append(["Bob", "bob@x.com", "Senior", "London", "Java", "2026-08-01", "high"])

    nj_ws = wb.create_sheet("New Joiners")
    nj_headers = ["Name", "Email", "Grade", "Location", "Key Skills", "Join Date"]
    nj_ws.append([""] * len(nj_headers))
    nj_ws.append(nj_headers)

    path = tmp_path / "consultants.xlsx"
    wb.save(path)
    return path


def test_ingest_roles_reads_sector_column(tmp_path: Path) -> None:
    path = _make_roles_workbook(tmp_path, with_sector=True)
    roles = ingest_roles(path)
    assert len(roles) == 1
    assert roles[0].sector == "FinTech"


def test_ingest_roles_sector_defaults_empty_when_missing(tmp_path: Path) -> None:
    path = _make_roles_workbook(tmp_path, with_sector=False)
    roles = ingest_roles(path)
    assert len(roles) == 1
    assert roles[0].sector == ""


def test_ingest_consultants_reads_days_on_beach(tmp_path: Path) -> None:
    path = _make_consultants_workbook(tmp_path, with_days_on_beach=True)
    consultants = ingest_consultants_from_workbook(path)
    beach = [c for c in consultants if c.supply_state == "beach"]
    assert len(beach) == 1
    assert beach[0].days_on_beach == 45


def test_ingest_consultants_days_on_beach_defaults_zero_when_missing(tmp_path: Path) -> None:
    path = _make_consultants_workbook(tmp_path, with_days_on_beach=False)
    consultants = ingest_consultants_from_workbook(path)
    beach = [c for c in consultants if c.supply_state == "beach"]
    assert len(beach) == 1
    assert beach[0].days_on_beach == 0


def test_ingest_consultants_rolling_off_days_on_beach_zero(tmp_path: Path) -> None:
    path = _make_consultants_workbook(tmp_path, with_days_on_beach=True, include_rolling_off=True)
    consultants = ingest_consultants_from_workbook(path)
    rolling_off = [c for c in consultants if c.supply_state == "rolling_off"]
    assert len(rolling_off) == 1
    assert rolling_off[0].days_on_beach == 0
