from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from matcher.models.errors import IngestionError
from matcher.pipeline.ingest import (
    ingest_consultants,
    ingest_consultants_from_workbook,
    ingest_roles,
)


def _make_workbook_with_roles_sheet(path: Path, sheet_name: str = "Open Roles") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["header row 1"])
    ws.append(["Role ID", "Title", "Required Skills", "Location"])
    ws.append(["ROLE-X", "Test Role", "Python", "London"])
    wb.save(path)


def test_zero_byte_workbook_raises_ingestion_error(tmp_path: Path) -> None:
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"")
    with pytest.raises(IngestionError, match="workbook unreadable"):
        ingest_roles(bad)


def test_missing_open_roles_sheet_raises_ingestion_error(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WrongSheet"
    path = tmp_path / "no_roles.xlsx"
    wb.save(path)
    with pytest.raises(IngestionError, match="missing sheet 'Open Roles'"):
        ingest_roles(path)


def test_missing_role_id_column_raises_ingestion_error(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Open Roles"
    ws.append([""])
    ws.append(["Title", "Required Skills", "Location"])
    ws.append(["Test Role", "Python", "London"])
    path = tmp_path / "missing_col.xlsx"
    wb.save(path)
    with pytest.raises(IngestionError, match="missing columns"):
        ingest_roles(path)


def test_corrupt_pdf_never_raises(tmp_path: Path) -> None:
    pdf_path = tmp_path / "corrupt_person_pp.pdf"
    pdf_path.write_bytes(b"not a valid pdf")

    from matcher.models.consultant import Consultant

    consultant = Consultant(name="Corrupt Person", email="corrupt@example.com")
    result = ingest_consultants(tmp_path, [consultant])
    assert len(result) == 1
    assert result[0].data_confidence < 1.0


def test_corrupt_pdf_gets_gap_flag(tmp_path: Path) -> None:
    pdf_path = tmp_path / "corrupt_person_pp.pdf"
    pdf_path.write_bytes(b"not a valid pdf")

    from matcher.models.consultant import Consultant

    consultant = Consultant(name="Corrupt Person", email="corrupt@example.com")
    result = ingest_consultants(tmp_path, [consultant])
    assert "profile_pdf_unreadable" in result[0].data_gaps


def test_missing_supply_sheet_raises_ingestion_error(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beach"
    ws.append([""])
    ws.append(["Name", "Email", "Grade", "Location"])
    path = tmp_path / "missing_supply_sheet.xlsx"
    wb.save(path)
    with pytest.raises(IngestionError, match="missing sheet"):
        ingest_consultants_from_workbook(path)
