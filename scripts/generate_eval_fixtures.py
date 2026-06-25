from __future__ import annotations

from pathlib import Path

import openpyxl

OUTPUT_PATH = Path("evals/fixtures/eval_data.xlsx")

_ROLES = [
    ("EVAL-01", "Senior Python Engineer", "Python (working); Django (working)", "London", ""),
    ("EVAL-02", "Java Developer", "Java (working); Spring (working)", "London", ""),
    ("EVAL-03", "C++ Embedded Engineer", "C++ (working); Embedded Systems (working)", "London", ""),
    (
        "EVAL-04",
        "COBOL Mainframe Engineer",
        "COBOL (expert); JCL (expert)",
        "Singapore",
        "",
    ),
    ("EVAL-05", "Python Data Engineer", "Python (working); SQL (working)", "London", ""),
]

_BEACH = [
    ("Alice Test", "alice.test@example.com", "Senior", "London", "Python, Django, FastAPI"),
    ("Bob Test", "bob.test@example.com", "Mid", "London", "Java, Spring, Maven"),
    ("Carol Test", "python.only@example.com", "Mid", "London", "Python, Django"),
    ("Dave Test", "highfeedback.test@example.com", "Mid", "London", "Python, SQL, Pandas"),
    ("Eve Test", "lowfeedback.test@example.com", "Mid", "London", "Python, SQL"),
]


def _write_roles_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Open Roles")
    ws.cell(1, 1, "Open Roles")
    headers = ["Role ID", "Title", "Required Skills", "Location", "Notes / Constraints"]
    for col, h in enumerate(headers, start=1):
        ws.cell(2, col, h)
    for row_idx, row_data in enumerate(_ROLES, start=3):
        for col, val in enumerate(row_data, start=1):
            ws.cell(row_idx, col, val)


def _write_consultant_sheet(
    wb: openpyxl.Workbook, sheet_name: str, rows: list[tuple[str, ...]]
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.cell(1, 1, sheet_name)
    headers = ["Name", "Email", "Grade", "Location", "Key Skills"]
    for col, h in enumerate(headers, start=1):
        ws.cell(2, col, h)
    for row_idx, row_data in enumerate(rows, start=3):
        for col, val in enumerate(row_data, start=1):
            ws.cell(row_idx, col, val)


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_roles_sheet(wb)
    _write_consultant_sheet(wb, "Beach", _BEACH)
    _write_consultant_sheet(wb, "Rolling Off", [])
    _write_consultant_sheet(wb, "New Joiners", [])

    wb.save(OUTPUT_PATH)
    print(f"Written: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
